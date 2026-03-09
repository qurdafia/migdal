import openpyxl
from io import BytesIO
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse

# --- ADDED TelemetryRecord HERE ---
from apps.core.models import DataSource, TelemetryRecord

# --- CSV & DATE IMPORTS ---
import csv
from django.utils.dateparse import parse_date
from django.utils import timezone
from datetime import datetime, timedelta

# --- PDF IMPORTS (ReportLab) ---
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors

from apps.ai_engine.models import AnalysisReport 
from reportlab.lib.utils import simpleSplit
from xhtml2pdf import pisa

from django.template.loader import render_to_string


class CategoryHealthReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, category, report_type):
        start_str = request.query_params.get('start_date')
        end_str = request.query_params.get('end_date')

        devices = DataSource.objects.filter(
            organization=request.user.organization,
            device_type__iexact=category
        )

        # --- HELPER: GET DATA SNAPSHOT ---
        def get_snapshot(device):
            # Handle Schema Fallback
            if hasattr(device, 'telemetry'):
                qs = device.telemetry.all()
            else:
                qs = device.telemetryrecord_set.all()
            
            # Robust Date Filtering (Fixed Timezone Warning)
            if start_str and end_str:
                try:
                    # Convert String -> Date -> Datetime (Midnight) -> Aware Datetime
                    s_date = parse_date(start_str)
                    e_date = parse_date(end_str)
                    
                    if s_date and e_date:
                        start_dt = timezone.make_aware(datetime.combine(s_date, datetime.min.time()))
                        end_dt = timezone.make_aware(datetime.combine(e_date, datetime.min.time())) + timedelta(days=1)
                        qs = qs.filter(timestamp__gte=start_dt, timestamp__lt=end_dt)
                except Exception as e:
                    print(f"Date Parse Error: {e}")
                    pass
            
            return qs.order_by('-timestamp').first()

        # --- HELPER: CALCULATE HEALTH ---
        def calculate_health(record):
            if not record or not record.payload:
                return "No Data", colors.gray

            # Extract CPU (Try common keys)
            # FIX: Use 'payload_data' instead of 'p' to avoid variable conflict
            payload_data = record.payload 
            cpu = payload_data.get('cpu_1min') or payload_data.get('cpu_usage') or payload_data.get('cluster_cpu_usage_pct') or 0
            mem = payload_data.get('memory_usage') or payload_data.get('ram_usage') or 0
            
            try: cpu = float(cpu)
            except: cpu = 0
            try: mem = float(mem)
            except: mem = 0

            if cpu > 90 or mem > 90:
                return "CRITICAL", colors.red
            if cpu > 75 or mem > 75:
                return "WARNING", colors.orange
            
            return "HEALTHY", colors.green

        # =================================================
        # CSV GENERATION
        # =================================================
        if report_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{category}_report.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'IP Address', 'Health Status', 'CPU Load', 'Memory', 'Last Timestamp'])
            
            for d in devices:
                rec = get_snapshot(d)
                health_status, _ = calculate_health(rec)
                
                cpu = "N/A"
                mem = "N/A"
                ip = "N/A"
                
                if rec and rec.payload:
                    payload_data = rec.payload
                    ip = payload_data.get('ip_address') or payload_data.get('host_ip') or "N/A"
                    cpu = payload_data.get('cpu_1min') or payload_data.get('cpu_usage') or payload_data.get('cluster_cpu_usage_pct') or "N/A"
                    mem = payload_data.get('memory_usage') or payload_data.get('ram_usage') or "N/A"

                writer.writerow([
                    d.name, 
                    ip, 
                    health_status,
                    f"{cpu}%" if cpu != "N/A" else "N/A",
                    f"{mem}%" if mem != "N/A" else "N/A",
                    timezone.localtime(rec.timestamp).strftime("%Y-%m-%d %H:%M:%S") if rec else "N/A"
                    # rec.timestamp.strftime("%Y-%m-%d %H:%M:%S") if rec else "N/A"
                ])
            return response

        # =================================================
        # PDF GENERATION
        # =================================================
        elif report_type == 'download':
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{category}_report.pdf"'

            # 'p' is the PDF Canvas
            p = canvas.Canvas(response, pagesize=letter)
            width, height = letter

            # Header
            p.setFont("Helvetica-Bold", 18)
            p.drawString(50, height - 50, f"Migdal Report: {category.upper()}")
            
            p.setFont("Helvetica", 10)
            date_range_text = f"Filter: {start_str} to {end_str}" if start_str else "Filter: All Time"
            p.drawString(50, height - 70, date_range_text)
            
            p.setLineWidth(1)
            p.line(50, height - 80, width - 50, height - 80)

            # Table Header
            y = height - 110
            p.setFont("Helvetica-Bold", 9)
            p.drawString(50, y, "DEVICE")
            p.drawString(180, y, "IP ADDR")
            p.drawString(300, y, "METRICS (CPU / MEM)")
            p.drawString(450, y, "HEALTH")
            
            y -= 10
            p.setLineWidth(0.5)
            p.line(50, y, width - 50, y)
            y -= 20

            # Rows
            p.setFont("Helvetica", 10)
            
            for d in devices:
                if y < 50: 
                    p.showPage()
                    y = height - 50
                    p.setFont("Helvetica", 10)

                rec = get_snapshot(d)
                health_text, health_color = calculate_health(rec)

                # Extract Data
                ip = "N/A"
                metrics_text = "-"
                
                if rec and rec.payload:
                    # FIX: Use 'payload_data' instead of 'p' so we don't overwrite the Canvas!
                    payload_data = rec.payload
                    ip = payload_data.get('ip_address') or payload_data.get('host_ip') or "N/A"
                    
                    c = payload_data.get('cpu_1min') or payload_data.get('cpu_usage') or payload_data.get('cpu_load') or payload_data.get('cluster_cpu_usage_pct') or "-"
                    m = payload_data.get('memory_usage') or payload_data.get('ram_usage') or "-"
                    
                    metrics_text = f"CPU: {c}%  |  MEM: {m}%"

                # Draw Name
                p.setFillColor(colors.black)
                p.drawString(50, y, d.name)
                
                # Draw IP
                p.setFillColor(colors.darkgray)
                p.setFont("Courier", 9)
                p.drawString(180, y, ip)
                
                # Draw Metrics
                p.setFont("Helvetica", 10)
                p.setFillColor(colors.black)
                p.drawString(300, y, metrics_text)
                
                # Draw Health Status (Color Coded)
                p.setFillColor(health_color)
                p.setFont("Helvetica-Bold", 10)
                p.drawString(450, y, health_text)

                y -= 20

            p.showPage()
            p.save()
            return response

        return Response({"error": "Invalid report type"}, status=400)


class DownloadReportView(APIView):
    """
    Generates a PDF using an HTML Template for AI Analysis.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request, report_id):
        # 1. Fetch Report
        try:
            report = AnalysisReport.objects.get(
                id=report_id, 
                source__organization=request.user.organization
            )
        except AnalysisReport.DoesNotExist:
            return Response({"error": "Report not found"}, status=404)

        # 2. Context for the Template
        # We need to manually fetch the IP because it's a calculated property on the model
        # or inside the latest telemetry payload.
        ip_address = "N/A"
        if hasattr(report.source, 'telemetry'):
             last = report.source.telemetry.first()
             if last and last.payload:
                 ip_address = last.payload.get('ip_address', 'N/A')

        context = {
            'report': report,
            'generated_at': timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S'),
            # 'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ip_address': ip_address
        }

        # 3. Render HTML
        html_string = render_to_string('reports/analysis_pdf.html', context)

        # 4. Convert to PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="AI_Analysis_{report.source.name}.pdf"'

        pisa_status = pisa.CreatePDF(
            html_string, 
            dest=response
        )

        if pisa_status.err:
            return Response({"error": "PDF generation failed"}, status=500)
            
        return response


# =====================================================================
# NEW: HISTORICAL REPORTING VIEW
# =====================================================================

class HistoricalReportView(APIView):
    """
    Generates a historical Excel (.xlsx) report separated into actual TABS by device type.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_str = request.query_params.get('start_date')
        end_str = request.query_params.get('end_date')

        try:
            s_date = parse_date(start_str)
            e_date = parse_date(end_str)
            start_dt = timezone.make_aware(datetime.combine(s_date, datetime.min.time()))
            end_dt = timezone.make_aware(datetime.combine(e_date, datetime.min.time())) + timedelta(days=1)
        except Exception:
            return HttpResponse("Invalid date format", status=400)

        # Query ALL records in the time range
        records = TelemetryRecord.objects.filter(
            source__organization=request.user.organization,
            timestamp__gte=start_dt,
            timestamp__lt=end_dt
        ).select_related('source').order_by('source__device_type', 'source__name', '-timestamp')

        # --- 1. INITIALIZE EXCEL WORKBOOK ---
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default empty sheet
        
        # Group the records into a dictionary by device_type
        grouped_data = {}
        for rec in records:
            dtype = str(rec.source.device_type).capitalize()
            if dtype not in grouped_data:
                grouped_data[dtype] = []
            grouped_data[dtype].append(rec)

        # Handle empty case
        if not grouped_data:
            ws = wb.create_sheet(title="No Data")
            ws.append(["No records found for this date range."])
        else:
            # --- 2. CREATE A TAB FOR EACH DEVICE TYPE ---
            for dtype, recs in grouped_data.items():
                ws = wb.create_sheet(title=dtype)
                
                # Create the Header Row
                ws.append(['Device Name', 'IP Address', 'Timestamp (Local)', 'CPU', 'Memory', 'Health Status'])
                
                # Format Header Row (Optional: Make it bold)
                for cell in ws["1:1"]:
                    cell.font = openpyxl.styles.Font(bold=True)

                for rec in recs:
                    payload_data = rec.payload or {}
                    ip = payload_data.get('ip_address') or payload_data.get('host_ip') or "N/A"
                    
                    c_val = payload_data.get('cpu_1min') or payload_data.get('cpu_usage') or payload_data.get('cluster_cpu_usage_pct') or 0
                    m_val = payload_data.get('memory_usage') or payload_data.get('ram_usage') or 0
                    
                    try: float_c = float(c_val)
                    except: float_c = 0
                    try: float_m = float(m_val)
                    except: float_m = 0

                    if float_c > 90 or float_m > 90: health = "CRITICAL"
                    elif float_c > 75 or float_m > 75: health = "WARNING"
                    else: health = "HEALTHY"

                    # --- 3. FIX THE TIMEZONE ---
                    # Converts the strict UTC database time to the user's local timezone
                    local_time = timezone.localtime(rec.timestamp)
                    formatted_time = local_time.strftime("%Y-%m-%d %H:%M:%S")

                    ws.append([
                        rec.source.name,
                        ip,
                        formatted_time,
                        f"{float_c}%" if float_c else "N/A",
                        f"{float_m}%" if float_m else "N/A",
                        health
                    ])

        # --- 4. SAVE AND RETURN EXCEL FILE ---
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Migdal_Historical_{start_str}_to_{end_str}.xlsx"'
        
        return response